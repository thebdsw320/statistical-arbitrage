from openpyxl import load_workbook
import pyautogui as pag
import os, time
import pandas as pd

def save_excel():
    os.system('wps backtest.xlsx &')
    time.sleep(7)
    pag.moveTo(800, 500)
    pag.click()
    pag.hotkey('ctrl', 's')
    time.sleep(3)
    pag.hotkey('alt', 'f4')

def update_data(pair_csv):
    wb = load_workbook(filename='./backtest.xlsx')
    sheet = wb['mean_reverting']
    pair_data = pd.read_csv(pair_csv)
    
    symbols = list(pair_data.columns)[1:3]
    
    sheet['B1'], sheet['C1'] = symbols[0], symbols[1]
    
    for i, v in pair_data[symbols[0]].iteritems():
        n = i+2
        sheet[f'B{n}'] = v
    
    for i, v in pair_data[symbols[1]].iteritems():
        n = i+2
        sheet[f'C{n}'] = v
        
    wb.save(filename='./backtest.xlsx')

def update_values(zscore_threshold, trading_capital, rebate, slippage_assumption, long_when_zscore_negative, many_opens):
    wb = load_workbook(filename='./backtest.xlsx')
    sheet = wb['mean_reverting']
    sheet['AH1'], sheet['AH2'], sheet['AH3'], sheet['AH4'], sheet['AH5'], sheet['AH6'] = float(zscore_threshold), float(trading_capital), float(rebate), float(slippage_assumption), long_when_zscore_negative, many_opens
    wb.save(filename='./backtest.xlsx')
    
def retrieve_data():
    save_excel()
    
    wb = load_workbook(filename='./backtest.xlsx', data_only=True)
    sheet = wb['mean_reverting']
    
    longs_profit, shorts_profit, net_profit = sheet['AH8'].value, sheet['AH9'].value, sheet['AH10'].value
    longs_profit_v, shorts_profit_v, net_profit_v = sheet['AI8'].value, sheet['AI9'].value, sheet['AI10'].value
    roi, roi_v = sheet['AH12'].value, sheet['AI12'].value
    win_rate_longs, win_rate_shorts, avg_win_rate = sheet['AH14'].value, sheet['AH15'].value, sheet['AH16'].value
    win_rate_longs_v, win_rate_shorts_v, avg_win_rate_v = sheet['AI14'].value, sheet['AI15'].value, sheet['AI16'].value
    best_long, best_short, worst_long, worst_short = sheet['AH18'].value, sheet['AH19'].value, sheet['AH20'].value, sheet['AH21'].value
    best_long_v, best_short_v, worst_long_v, worst_short_v = sheet['AI18'].value, sheet['AI19'].value, sheet['AI20'].value, sheet['AI21'].value
    
    headings = ['Description', 'Value']
    data = [
        [longs_profit, round(longs_profit_v, 2)],
        [shorts_profit, round(shorts_profit_v, 2)],
        [net_profit, round(net_profit_v, 2)],
        [roi, f'{round(roi_v*100, 2)}%'],
        [win_rate_longs, f'{round(win_rate_longs_v*100, 2)}%'],
        [win_rate_shorts, f'{round(win_rate_shorts_v*100, 2)}%'],
        [avg_win_rate, f'{round(avg_win_rate_v*100, 2)}%'],
        [best_long, f'{round(best_long_v*100, 2)}%'],
        [best_short, f'{round(best_short_v*100, 2)}%'],
        [worst_long, f'{round(worst_long_v*100, 2)}%'],
        [worst_short, f'{round(worst_short_v*100, 2)}%']
    ]
    
    return headings, data